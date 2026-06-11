import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import openpyxl

st.set_page_config(
    page_title="Madrugón — Descuentos",
    page_icon="🏷️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============ ESTILOS ============
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Reset Streamlit defaults */
.main .block-container {
    padding: 0 !important;
    max-width: 100% !important;
}
section[data-testid="stSidebar"] { display: none; }
header[data-testid="stHeader"] { display: none; }
footer { display: none; }

/* ---- LAYOUT SHELL ---- */
.app-shell {
    min-height: 100vh;
    background: #F5F6FA;
    display: flex;
    flex-direction: column;
}

/* ---- TOP BAR ---- */
.topbar {
    background: #0F172A;
    padding: 0 32px;
    height: 64px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    position: sticky;
    top: 0;
    z-index: 100;
}
.topbar-brand {
    display: flex;
    align-items: center;
    gap: 12px;
}
.topbar-logo {
    width: 36px;
    height: 36px;
    background: linear-gradient(135deg, #FF5F57 0%, #FF8C42 100%);
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 18px;
}
.topbar-name {
    color: white;
    font-weight: 700;
    font-size: 18px;
    letter-spacing: -0.3px;
}
.topbar-sub {
    color: #64748B;
    font-size: 13px;
    font-weight: 400;
}
.topbar-badge {
    background: #1E293B;
    color: #94A3B8;
    padding: 6px 14px;
    border-radius: 20px;
    font-size: 13px;
    font-weight: 500;
}
.topbar-badge span {
    color: #38BDF8;
    font-weight: 700;
}

/* ---- UPLOAD SCREEN ---- */
.upload-screen {
    flex: 1;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 40px 20px;
}
.upload-card {
    background: white;
    border-radius: 24px;
    padding: 60px 48px;
    text-align: center;
    max-width: 480px;
    width: 100%;
    box-shadow: 0 4px 24px rgba(0,0,0,0.06);
    border: 1px solid #E2E8F0;
}
.upload-icon {
    width: 80px;
    height: 80px;
    background: linear-gradient(135deg, #EFF6FF, #DBEAFE);
    border-radius: 20px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 36px;
    margin: 0 auto 24px;
}
.upload-title {
    font-size: 24px;
    font-weight: 700;
    color: #0F172A;
    margin-bottom: 10px;
}
.upload-sub {
    color: #64748B;
    font-size: 15px;
    line-height: 1.6;
    margin-bottom: 32px;
}

/* ---- SEARCH SCREEN ---- */
.search-screen {
    flex: 1;
    display: flex;
    flex-direction: column;
}
.search-hero {
    background: linear-gradient(135deg, #0F172A 0%, #1E3A5F 100%);
    padding: 40px 32px 48px;
}
.search-hero-title {
    color: white;
    font-size: 28px;
    font-weight: 800;
    letter-spacing: -0.5px;
    margin-bottom: 6px;
}
.search-hero-sub {
    color: #94A3B8;
    font-size: 15px;
    margin-bottom: 28px;
}
.search-box-wrap {
    position: relative;
    max-width: 640px;
}
.search-icon-abs {
    position: absolute;
    left: 18px;
    top: 50%;
    transform: translateY(-50%);
    font-size: 18px;
    z-index: 1;
    pointer-events: none;
}

/* Override Streamlit input */
.search-box-wrap .stTextInput > div > div > input {
    background: white !important;
    border: none !important;
    border-radius: 14px !important;
    padding: 16px 20px 16px 52px !important;
    font-size: 16px !important;
    font-weight: 500 !important;
    color: #0F172A !important;
    height: 56px !important;
    box-shadow: 0 8px 32px rgba(0,0,0,0.2) !important;
}
.search-box-wrap .stTextInput > div > div > input:focus {
    box-shadow: 0 8px 32px rgba(0,0,0,0.2), 0 0 0 3px rgba(56,189,248,0.4) !important;
    outline: none !important;
}
.search-box-wrap .stTextInput label { display: none !important; }
.search-box-wrap .stTextInput > div { padding: 0 !important; }

/* ---- RESULTS AREA ---- */
.results-area {
    padding: 32px;
    flex: 1;
}

/* ---- EMPTY STATE ---- */
.empty-state {
    text-align: center;
    padding: 60px 20px;
    color: #94A3B8;
}
.empty-state-icon { font-size: 48px; margin-bottom: 16px; }
.empty-state-title { font-size: 18px; font-weight: 600; color: #64748B; margin-bottom: 8px; }
.empty-state-sub { font-size: 14px; }

/* ---- NOT FOUND ---- */
.not-found {
    background: #FFF1F2;
    border: 1px solid #FFE4E6;
    border-radius: 16px;
    padding: 32px;
    text-align: center;
    max-width: 480px;
    margin: 0 auto;
}
.not-found-icon { font-size: 40px; margin-bottom: 12px; }
.not-found-title { font-size: 18px; font-weight: 700; color: #BE123C; margin-bottom: 8px; }
.not-found-sub { color: #64748B; font-size: 14px; }

/* ---- PRODUCT RESULT ---- */
.result-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
    max-width: 960px;
    margin: 0 auto;
}
@media (max-width: 700px) {
    .result-grid { grid-template-columns: 1fr; }
    .results-area { padding: 16px; }
    .search-hero { padding: 24px 16px 32px; }
}

/* MAIN DISCOUNT CARD */
.discount-card {
    background: linear-gradient(145deg, #FF5F57 0%, #FF8C42 100%);
    border-radius: 20px;
    padding: 32px;
    color: white;
    position: relative;
    overflow: hidden;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    min-height: 220px;
}
.discount-card::before {
    content: '';
    position: absolute;
    top: -30px; right: -30px;
    width: 160px; height: 160px;
    background: rgba(255,255,255,0.08);
    border-radius: 50%;
}
.discount-card::after {
    content: '';
    position: absolute;
    bottom: -50px; left: -20px;
    width: 200px; height: 200px;
    background: rgba(255,255,255,0.05);
    border-radius: 50%;
}
.dc-eyebrow {
    font-size: 12px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    opacity: 0.8;
}
.dc-pct {
    font-size: 88px;
    font-weight: 900;
    line-height: 1;
    letter-spacing: -4px;
    position: relative;
    z-index: 1;
}
.dc-pct sup {
    font-size: 32px;
    letter-spacing: 0;
    vertical-align: super;
}
.dc-label {
    font-size: 18px;
    font-weight: 700;
    opacity: 0.9;
}
.dc-savings {
    background: rgba(255,255,255,0.2);
    border-radius: 10px;
    padding: 10px 16px;
    font-size: 14px;
    font-weight: 600;
    display: inline-block;
    position: relative;
    z-index: 1;
    margin-top: 12px;
}

/* No discount */
.no-discount-card {
    background: #F1F5F9;
    border-radius: 20px;
    padding: 32px;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    text-align: center;
    min-height: 220px;
    color: #94A3B8;
}
.no-discount-icon { font-size: 40px; margin-bottom: 12px; }
.no-discount-text { font-size: 16px; font-weight: 600; color: #64748B; }

/* PRODUCT INFO CARD */
.info-card {
    background: white;
    border-radius: 20px;
    padding: 28px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.05);
    border: 1px solid #F1F5F9;
    display: flex;
    flex-direction: column;
    gap: 20px;
}
.product-name {
    font-size: 20px;
    font-weight: 700;
    color: #0F172A;
    line-height: 1.3;
}
.product-brand {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 13px;
    font-weight: 500;
    color: #475569;
}
.price-row {
    display: flex;
    gap: 16px;
    align-items: flex-end;
}
.price-block { flex: 1; }
.price-label {
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #94A3B8;
    margin-bottom: 4px;
}
.price-original {
    font-size: 22px;
    font-weight: 600;
    color: #CBD5E1;
    text-decoration: line-through;
}
.price-discounted {
    font-size: 32px;
    font-weight: 800;
    color: #FF5F57;
    letter-spacing: -1px;
}
.price-same {
    font-size: 28px;
    font-weight: 800;
    color: #0F172A;
    letter-spacing: -1px;
}
.divider { height: 1px; background: #F1F5F9; }
.meta-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 12px;
}
.meta-item {
    background: #F8FAFC;
    border-radius: 10px;
    padding: 10px 14px;
}
.meta-label {
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    color: #94A3B8;
    margin-bottom: 3px;
}
.meta-value {
    font-size: 14px;
    font-weight: 600;
    color: #334155;
    word-break: break-all;
}
.status-pill {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 12px;
    font-weight: 600;
}
.status-activo { background: #DCFCE7; color: #16A34A; }
.status-pasivo { background: #FEF2F2; color: #DC2626; }

/* DEBUG SECTION */
.debug-wrap {
    margin-top: 24px;
    max-width: 960px;
    margin-left: auto;
    margin-right: auto;
}

/* Streamlit expander override */
.streamlit-expanderHeader {
    font-size: 13px !important;
    color: #64748B !important;
}

/* Stats bar */
.stats-bar {
    background: #1E293B;
    padding: 10px 32px;
    display: flex;
    gap: 24px;
    flex-wrap: wrap;
}
.stat-item {
    display: flex;
    align-items: center;
    gap: 8px;
}
.stat-dot {
    width: 8px; height: 8px;
    border-radius: 50%;
}
.stat-dot-green { background: #22C55E; }
.stat-dot-blue { background: #38BDF8; }
.stat-text {
    font-size: 13px;
    color: #64748B;
}
.stat-text span { color: #E2E8F0; font-weight: 600; }
</style>
""", unsafe_allow_html=True)


# ============ ESTADO ============
if 'df' not in st.session_state:
    st.session_state.df = None
if 'meta' not in st.session_state:
    st.session_state.meta = {}
if 'debug' not in st.session_state:
    st.session_state.debug = False


# ============ CARGA DE DATOS ============
@st.cache_data
def cargar_excel(archivo):
    wb = openpyxl.load_workbook(archivo, read_only=True)
    hoja = wb.sheetnames[0]
    # Leer fila 0 para detectar días de descuento
    ws = wb[hoja]
    fila0 = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    dias_descuento = [v for v in fila0 if v and isinstance(v, str) and 'DESCUENTO' in v.upper()]
    wb.close()

    df = pd.read_excel(archivo, sheet_name=hoja, header=1)
    df.columns = [str(c).strip() for c in df.columns]

    # Renombrar columnas duplicadas (PVP, DCTO, PVP CON DESUENTO aparecen dos veces)
    seen = {}
    new_cols = []
    for col in df.columns:
        if col in seen:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            new_cols.append(col)
    df.columns = new_cols

    # Convertir EAN a string
    for col in df.columns:
        if 'EAN' in col.upper() or 'CODIGO' in col.upper():
            df[col] = df[col].astype(str).str.strip().str.split('.').str[0]

    total = len(df)
    activos = int((df['ESTADO'].str.upper() == 'ACTIVO').sum()) if 'ESTADO' in df.columns else 0
    meta = {
        'hoja': hoja,
        'total': total,
        'activos': activos,
        'dias': dias_descuento
    }
    return df, meta


def buscar_producto(df, codigo):
    mascara = pd.Series(False, index=df.index)
    for col in df.columns:
        cu = col.upper()
        if any(k in cu for k in ['EAN', 'CODIGO', 'SAP']):
            try:
                mascara |= (df[col].astype(str).str.strip() == codigo.strip())
            except:
                pass
    return df[mascara]


def extraer_info(fila, df_cols):
    cols = {c.upper(): c for c in df_cols}

    def v(key_parts, exclude=None):
        for k, real in cols.items():
            if any(p in k for p in key_parts):
                if exclude and any(e in k for e in exclude):
                    continue
                val = fila[real]
                if not pd.isna(val) and str(val).strip() not in ('', 'nan'):
                    return val
        return None

    nombre = v(['NOMBRE'], exclude=['PROVEEDOR'])
    marca  = v(['PROVEEDOR', 'MARCA', 'BRAND'], exclude=['NIT', 'NUMERO', 'NUMBER'])
    estado = v(['ESTADO'])
    ean    = fila.get(cols.get('EAN', ''), None)
    sap    = fila.get(cols.get('CODIGO SAP', ''), None)

    # Precios: tomar PVP base (primer PVP sin sufijo) y PVP CON DESUENTO
    pvp_base = None
    pvp_desc = None
    dcto_pct = None

    for c in df_cols:
        cu = c.upper()
        val = fila[c]
        if pd.isna(val): continue
        try:
            n = float(str(val).replace(',', '.').replace('%', '').strip())
        except:
            continue

        if cu == 'PVP' and pvp_base is None and n > 100:
            pvp_base = n
        if 'PVP CON' in cu and pvp_desc is None and n > 100:
            pvp_desc = n
        if 'DCTO' in cu and dcto_pct is None and 0 < n <= 1:
            dcto_pct = round(n * 100, 1)

    # Calcular descuento si falta
    if dcto_pct is None and pvp_base and pvp_desc and pvp_base > pvp_desc:
        dcto_pct = round((pvp_base - pvp_desc) / pvp_base * 100, 1)

    if pvp_desc is None and pvp_base and dcto_pct:
        pvp_desc = round(pvp_base * (1 - dcto_pct / 100), 0)

    return {
        'nombre': str(nombre).strip() if nombre else 'Sin nombre',
        'marca':  str(marca).strip() if marca else None,
        'estado': str(estado).strip().upper() if estado else None,
        'ean':    str(ean).split('.')[0] if ean else None,
        'sap':    str(sap).split('.')[0] if sap else None,
        'pvp':    pvp_base,
        'pvp_desc': pvp_desc,
        'dcto':   dcto_pct,
        'ahorro': round(pvp_base - pvp_desc, 0) if pvp_base and pvp_desc else None
    }


# ============ TOP BAR ============
df = st.session_state.df
meta = st.session_state.meta

loaded = df is not None

badge_html = ""
if loaded:
    badge_html = f'<div class="topbar-badge">📦 <span>{meta["total"]:,}</span> productos cargados</div>'
else:
    badge_html = '<div class="topbar-badge">Sin datos</div>'

st.markdown(f"""
<div class="topbar">
  <div class="topbar-brand">
    <div class="topbar-logo">🏷️</div>
    <div>
      <div class="topbar-name">Madrugón</div>
      <div class="topbar-sub">Consulta de descuentos</div>
    </div>
  </div>
  {badge_html}
</div>
""", unsafe_allow_html=True)


# ============ STATS BAR (cuando hay datos) ============
if loaded:
    dias_str = ' · '.join(meta['dias']) if meta['dias'] else 'Sin días definidos'
    st.markdown(f"""
    <div class="stats-bar">
      <div class="stat-item">
        <div class="stat-dot stat-dot-green"></div>
        <div class="stat-text"><span>{meta['activos']:,}</span> activos</div>
      </div>
      <div class="stat-item">
        <div class="stat-dot stat-dot-blue"></div>
        <div class="stat-text"><span>{meta['total'] - meta['activos']:,}</span> pasivos</div>
      </div>
      <div class="stat-item">
        <div class="stat-text">📅 {dias_str}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)


# ============ PANTALLA: SIN DATOS ============
if not loaded:
    st.markdown('<div class="upload-screen">', unsafe_allow_html=True)
    st.markdown("""
    <div class="upload-card">
      <div class="upload-icon">📊</div>
      <div class="upload-title">Cargar archivo de datos</div>
      <div class="upload-sub">Sube el archivo Excel del Madrugón para comenzar a consultar descuentos por código de barras.</div>
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("", type=["xlsx", "xls"], label_visibility="collapsed")

    if uploaded:
        with st.spinner("Procesando archivo..."):
            try:
                df_loaded, meta_loaded = cargar_excel(uploaded)
                st.session_state.df = df_loaded
                st.session_state.meta = meta_loaded
                st.rerun()
            except Exception as e:
                st.error(f"Error al cargar el archivo: {e}")

    st.markdown('</div>', unsafe_allow_html=True)


# ============ PANTALLA: BÚSQUEDA ============
else:
    # Hero + input
    st.markdown("""
    <div class="search-hero">
      <div class="search-hero-title">¿Qué producto buscas?</div>
      <div class="search-hero-sub">Escanea o escribe el código EAN o código SAP del producto</div>
    </div>
    """, unsafe_allow_html=True)

    # Input field — dentro de wrapper para CSS
    col_inp, col_act = st.columns([6, 1])
    with col_inp:
        st.markdown('<div class="search-box-wrap"><span class="search-icon-abs">🔍</span>', unsafe_allow_html=True)
        codigo = st.text_input("Código", key="codigo", label_visibility="collapsed",
                               placeholder="Escanea o escribe el código de barras...")
        st.markdown('</div>', unsafe_allow_html=True)
    with col_act:
        limpiar = st.button("🗑 Nuevo", use_container_width=True)
        if limpiar:
            st.session_state.df = None
            st.session_state.meta = {}
            st.rerun()

    # ---- RESULTADOS ----
    st.markdown('<div class="results-area">', unsafe_allow_html=True)

    if not codigo:
        st.markdown("""
        <div class="empty-state">
          <div class="empty-state-icon">🏷️</div>
          <div class="empty-state-title">Listo para buscar</div>
          <div class="empty-state-sub">Ingresa un código para ver el precio y descuento del producto</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        resultado = buscar_producto(df, codigo)

        if resultado.empty:
            st.markdown(f"""
            <div class="not-found">
              <div class="not-found-icon">🔎</div>
              <div class="not-found-title">Producto no encontrado</div>
              <div class="not-found-sub">No hay resultados para el código <strong>{codigo}</strong>.<br>Verifica que el código sea correcto e inténtalo de nuevo.</div>
            </div>
            """, unsafe_allow_html=True)
        else:
            fila = resultado.iloc[0]
            info = extraer_info(fila, df.columns.tolist())

            # ---- TARJETA DESCUENTO ----
            if info['dcto'] and info['dcto'] > 0:
                ahorro_html = ""
                if info['ahorro']:
                    ahorro_html = f'<div class="dc-savings">💰 Ahorras ${info["ahorro"]:,.0f}</div>'
                discount_card = f"""
                <div class="discount-card">
                  <div>
                    <div class="dc-eyebrow">🔥 Descuento especial</div>
                    <div class="dc-pct"><sup>-</sup>{info['dcto']:.0f}<sup>%</sup></div>
                    <div class="dc-label">DE DESCUENTO</div>
                  </div>
                  {ahorro_html}
                </div>
                """
            else:
                discount_card = """
                <div class="no-discount-card">
                  <div class="no-discount-icon">🏷️</div>
                  <div class="no-discount-text">Sin descuento este día</div>
                </div>
                """

            # ---- TARJETA PRODUCTO ----
            # Nombre + marca
            marca_html = f'<div><span class="product-brand">🏭 {info["marca"]}</span></div>' if info['marca'] else ''

            # Estado
            if info['estado']:
                cls = 'status-activo' if info['estado'] == 'ACTIVO' else 'status-pasivo'
                estado_html = f'<span class="status-pill {cls}">{info["estado"]}</span>'
            else:
                estado_html = ''

            # Precios
            if info['pvp'] and info['pvp_desc'] and info['pvp'] != info['pvp_desc']:
                precios_html = f"""
                <div class="price-row">
                  <div class="price-block">
                    <div class="price-label">Precio original</div>
                    <div class="price-original">${info['pvp']:,.0f}</div>
                  </div>
                  <div class="price-block">
                    <div class="price-label">Precio con descuento</div>
                    <div class="price-discounted">${info['pvp_desc']:,.0f}</div>
                  </div>
                </div>
                """
            elif info['pvp']:
                precios_html = f"""
                <div class="price-row">
                  <div class="price-block">
                    <div class="price-label">Precio de venta</div>
                    <div class="price-same">${info['pvp']:,.0f}</div>
                  </div>
                </div>
                """
            else:
                precios_html = '<div style="color:#94A3B8;font-size:14px;">Precio no disponible</div>'

            # Meta
            ean_val = info['ean'] or '—'
            sap_val = info['sap'] or '—'
            meta_html = f"""
            <div class="meta-grid">
              <div class="meta-item">
                <div class="meta-label">EAN</div>
                <div class="meta-value">{ean_val}</div>
              </div>
              <div class="meta-item">
                <div class="meta-label">SAP</div>
                <div class="meta-value">{sap_val}</div>
              </div>
            </div>
            """

            info_card = f"""
            <div class="info-card">
              <div>
                <div class="product-name">{info['nombre']}</div>
                {marca_html}
                {estado_html}
              </div>
              <div class="divider"></div>
              {precios_html}
              <div class="divider"></div>
              {meta_html}
            </div>
            """

            # Render grid usando components.html para evitar sanitización de Streamlit
            card_css = """
            <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
            * { box-sizing: border-box; margin: 0; padding: 0; font-family: 'Inter', sans-serif; }
            body { background: transparent; }
            .result-grid {
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
                padding: 4px;
            }
            @media (max-width: 600px) { .result-grid { grid-template-columns: 1fr; } }
            .discount-card {
                background: linear-gradient(145deg, #FF5F57 0%, #FF8C42 100%);
                border-radius: 20px;
                padding: 32px;
                color: white;
                position: relative;
                overflow: hidden;
                display: flex;
                flex-direction: column;
                justify-content: space-between;
                min-height: 220px;
            }
            .discount-card::before {
                content: '';
                position: absolute;
                top: -30px; right: -30px;
                width: 160px; height: 160px;
                background: rgba(255,255,255,0.08);
                border-radius: 50%;
            }
            .discount-card::after {
                content: '';
                position: absolute;
                bottom: -50px; left: -20px;
                width: 200px; height: 200px;
                background: rgba(255,255,255,0.05);
                border-radius: 50%;
            }
            .dc-eyebrow { font-size: 12px; font-weight: 600; text-transform: uppercase; letter-spacing: 1.5px; opacity: 0.8; }
            .dc-pct { font-size: 88px; font-weight: 900; line-height: 1; letter-spacing: -4px; position: relative; z-index: 1; }
            .dc-pct sup { font-size: 32px; letter-spacing: 0; vertical-align: super; }
            .dc-label { font-size: 18px; font-weight: 700; opacity: 0.9; }
            .dc-savings {
                background: rgba(255,255,255,0.2);
                border-radius: 10px;
                padding: 10px 16px;
                font-size: 14px;
                font-weight: 600;
                display: inline-block;
                position: relative;
                z-index: 1;
                margin-top: 12px;
            }
            .no-discount-card {
                background: #F1F5F9;
                border-radius: 20px;
                padding: 32px;
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                text-align: center;
                min-height: 220px;
                color: #94A3B8;
            }
            .no-discount-icon { font-size: 40px; margin-bottom: 12px; }
            .no-discount-text { font-size: 16px; font-weight: 600; color: #64748B; }
            .info-card {
                background: white;
                border-radius: 20px;
                padding: 28px;
                box-shadow: 0 2px 12px rgba(0,0,0,0.05);
                border: 1px solid #F1F5F9;
                display: flex;
                flex-direction: column;
                gap: 18px;
            }
            .product-name { font-size: 18px; font-weight: 700; color: #0F172A; line-height: 1.35; }
            .product-brand {
                display: inline-flex; align-items: center; gap: 6px;
                background: #F8FAFC; border: 1px solid #E2E8F0;
                padding: 4px 12px; border-radius: 20px;
                font-size: 13px; font-weight: 500; color: #475569;
                margin-top: 8px;
            }
            .status-pill { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; margin-top: 6px; }
            .status-activo { background: #DCFCE7; color: #16A34A; }
            .status-pasivo { background: #FEF2F2; color: #DC2626; }
            .divider { height: 1px; background: #F1F5F9; }
            .price-row { display: flex; gap: 16px; align-items: flex-start; }
            .price-block { flex: 1; }
            .price-label { font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 1px; color: #94A3B8; margin-bottom: 4px; }
            .price-original { font-size: 20px; font-weight: 600; color: #CBD5E1; text-decoration: line-through; }
            .price-discounted { font-size: 30px; font-weight: 800; color: #FF5F57; letter-spacing: -1px; }
            .price-same { font-size: 28px; font-weight: 800; color: #0F172A; letter-spacing: -1px; }
            .meta-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }
            .meta-item { background: #F8FAFC; border-radius: 10px; padding: 10px 14px; }
            .meta-label { font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.8px; color: #94A3B8; margin-bottom: 3px; }
            .meta-value { font-size: 14px; font-weight: 600; color: #334155; word-break: break-all; }
            </style>
            """
            components.html(f"""
            {card_css}
            <div class="result-grid">
              {discount_card}
              {info_card}
            </div>
            """, height=300, scrolling=False)

            # Múltiples resultados
            if len(resultado) > 1:
                with st.expander(f"⚠️ {len(resultado)} coincidencias encontradas — mostrando la primera"):
                    st.dataframe(resultado, use_container_width=True, hide_index=True)

            # Debug
            debug_col, _ = st.columns([1, 3])
            with debug_col:
                st.session_state.debug = st.checkbox("Modo diagnóstico", value=st.session_state.debug)

            if st.session_state.debug:
                st.markdown('<div class="debug-wrap">', unsafe_allow_html=True)
                with st.expander("🔧 Datos técnicos"):
                    st.write("**Información extraída:**")
                    st.json(info)
                    st.write("**Fila completa:**")
                    st.dataframe(pd.DataFrame([fila]), use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)
